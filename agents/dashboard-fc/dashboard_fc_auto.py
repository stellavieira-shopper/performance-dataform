"""
Dashboard FC — Geração automática de mensagens de performance semanal.

Uso:
  python dashboard_fc_auto.py --data DD/MM/AAAA

Variáveis de ambiente obrigatórias:
  SHEETS_TOKEN_PATH  — caminho para o sheets_token.json (OAuth2 pessoal)
                       Em Actions: gravado de SHEETS_TOKEN_JSON secret.
  ANTHROPIC_API_KEY  — chave da API Anthropic
"""

import argparse
import base64
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta

import google.generativeai as genai
import gspread
import openpyxl
import requests
from dotenv import load_dotenv
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ── IDs das planilhas ────────────────────────────────────────────────────────
SPREADSHEET_ID_FC   = "1j-mRJN2IHR3LjHaFtVgHtg2pH1XS3becphjBVjD02Ug"
SPREADSHEET_ID_GE   = "1E9ZXe2LtYySON6YYZYpn3BFEk-RiU-NsBJzmg6FUqrw"
EXPORT_MIME         = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TOKEN_PATH = os.getenv("SHEETS_TOKEN_PATH") or os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "sheets_token.json"
)

# Cores
COR_VERDE_ESCURO    = "#274e13"
COR_LARANJA         = "#b45309"

# ── Auth ─────────────────────────────────────────────────────────────────────

def get_credentials():
    creds = Credentials.from_authorized_user_file(TOKEN_PATH)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return creds


def download_xlsx(file_id: str, creds) -> openpyxl.Workbook:
    token = creds.token
    if not token:
        import google.auth.transport.requests
        creds.refresh(google.auth.transport.requests.Request())
        token = creds.token
    url = (
        f"https://www.googleapis.com/drive/v3/files/{file_id}/export"
        f"?mimeType={EXPORT_MIME}"
    )
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content))


# ── Parsing ───────────────────────────────────────────────────────────────────

def extract_val(cell_val):
    """Extrai valor real de células com fórmula DUMMYFUNCTION do IMPORTRANGE."""
    if cell_val is None:
        return None
    s = str(cell_val)
    m = re.search(r'DUMMYFUNCTION\("""COMPUTED_VALUE"""\),(.+)\)$', s)
    if m:
        v = m.group(1).strip()
        if v.startswith('"') and v.endswith('"'):
            v = v[1:-1]
        try:
            return float(v)
        except ValueError:
            return v
    return cell_val


def ler_resumo_fc(wb, fc: str) -> list[dict]:
    ws = wb[f"Resumo {fc}"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    return [dict(zip(header, row)) for row in rows[1:] if any(v is not None for v in row)]


def ler_visao_fc(wb, fc: str) -> list[dict]:
    ws = wb[f"Visão {fc}"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    return [
        dict(zip(header, row))
        for row in rows[1:]
        if row[0]  # SETOR preenchido
    ]


def ler_kpis_individuais(wb, data_quinta: str) -> list[dict]:
    """Lê aba 'KPIs Individuais' filtrando pela DATA (QUINTA) da semana."""
    ws = wb["KPIs Individuais"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]

    # data_quinta no formato DD/MM/YYYY
    try:
        target_date = datetime.strptime(data_quinta, "%d/%m/%Y").date()
    except ValueError:
        target_date = None

    result = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        d = dict(zip(header, row))
        if not d.get("DATA (QUINTA)"):
            continue

        cell_date = d["DATA (QUINTA)"]
        if hasattr(cell_date, "date"):
            row_date = cell_date.date()
        else:
            try:
                row_date = datetime.strptime(str(cell_date), "%d/%m/%Y").date()
            except ValueError:
                continue

        if target_date is None or row_date == target_date:
            d["DATA (QUINTA)"] = row_date.strftime("%d/%m/%Y") if row_date else str(cell_date)
            result.append(d)

    return result


def ler_vistoria_semana(wb) -> list[dict]:
    ws = wb["Vistoria Semana"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    result = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        d = dict(zip(header, row))
        if d.get("MATRÍCULA") is None:
            continue
        result.append(d)
    return result


def ler_ge(wb_ge) -> list[dict]:
    """Lê abas FC1/FC2/FC3 da planilha GE, retorna linhas com JUSTIFICATIVA."""
    result = []
    for fc_nome in ["FC1", "FC2", "FC3"]:
        if fc_nome not in wb_ge.sheetnames:
            continue
        ws = wb_ge[fc_nome]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        for row in rows[1:]:
            if not any(v is not None for v in row):
                continue
            matricula = row[0]
            justificativa = row[8] if len(row) > 8 else None
            if not justificativa:
                continue
            mat_str = None
            if matricula is not None:
                try:
                    mat_str = str(int(float(matricula)))
                except (ValueError, TypeError):
                    mat_str = str(matricula)
            result.append({
                "matricula": mat_str,
                "fc": fc_nome,
                "justificativa": str(justificativa),
            })
    return result


# ── Geração de mensagens com Anthropic ───────────────────────────────────────

def gerar_mensagem_setor(client, fc: str, config: dict, resumo: list[dict]) -> str:
    setor = config.get("SETOR", "")
    atribuicao = config.get("ATRIBUIÇÃO", "") or ""
    turno = config.get("TURNO", "") or ""
    desconto = config.get("DESCONTO", "") or ""
    ideia_central = config.get("IDEIA CENTRAL", "") or ""

    kpis_str = json.dumps(resumo, ensure_ascii=False, default=str)

    prompt = f"""Você é um assistente que gera mensagens de performance para colaboradores de um centro de distribuição (FC).

FC: {fc}
Setor: {setor}
Atribuição: {atribuicao or "TODAS"}
Turno: {turno or "TODOS"}
Desconto: {desconto or "sem desconto"}
Ideia central: {ideia_central}

KPIs da semana (Resumo {fc}):
{kpis_str}

Gere uma mensagem de performance seguindo EXATAMENTE estas regras:
1. Máximo 3 frases. Sem saudação. Sem assinatura. Sem emoji. Sem markdown.
2. Se há desconto (ex: -25%): primeira frase SEMPRE "{desconto} na Bonificação."
3. Se desconto = 0% por erro/processo especial: começar com "VALOR DA BONIFICAÇÃO ZERADO." e explicar o motivo vindo da ideia central.
4. Se desconto = 0% sem situação especial: começar direto pelo resultado.
5. Citar nomes reais dos indicadores (SMD, leva A, HR, Fiorino, pedidos mapeados, completos fresh/mercearia, rupturas, divergências).
6. NUNCA mencionar números, percentuais ou valores — apenas direção (melhorou, piorou, distante da meta).
7. Tom direto e factual — sem motivacionais genéricos.
8. Se turno específico (não TODOS): mencionar o turno.
9. Use a ideia central como fio condutor — mesmo que os dados não mostrem explicitamente.

Retorne APENAS o texto da mensagem, sem aspas ou explicações."""

    resp = client.generate_content(prompt)
    return resp.text.strip()


def gerar_mensagem_ge(client, item: dict) -> str:
    justificativa = item["justificativa"]

    prompt = f"""Você gera mensagens para colaboradores de inventário de Gestão de Estoque (GE).

Justificativa da planilha (contexto interno, não copiar diretamente):
{justificativa}

Gere uma mensagem reformulada seguindo EXATAMENTE estas regras:
1. Tom profissional e direto. Sem saudação. Sem assinatura. Sem emoji. Sem markdown.
2. NUNCA incluir números, percentuais ou valores — apenas direção qualitativa.
3. Use os padrões abaixo conforme o caso identificado na justificativa:
   - Não atingiu mínimo de posições: "Você não atingiu o mínimo de posições necessário para pontuar pela atividade de inventário de Gestão de Estoque nesta semana. Valide junto às suas lideranças dentro de Gestão de Estoque."
   - Não atingiu acurácia mínima: "Você não atingiu a acuracidade mínima necessária para pontuar pela atividade de inventário de Gestão de Estoque nesta semana. Valide junto às suas lideranças dentro de Gestão de Estoque."
   - Não atingiu posições E acurácia: "Você não atingiu o mínimo de posições nem a acuracidade mínima necessários para pontuar pela atividade de inventário de Gestão de Estoque nesta semana. Valide junto às suas lideranças dentro de Gestão de Estoque."
   - Detratora: "Você recebeu uma detratora pela atividade de inventário de Gestão de Estoque nesta semana devido ao baixo desempenho nas contagens. Precisamos reduzir os erros para conseguir pontuar positivamente por essa atividade e ficar mais próximo da bonificação."
4. Adapte o padrão acima se a justificativa apresentar uma situação diferente, mantendo o estilo e sem mencionar valores.

Retorne APENAS o texto da mensagem."""

    resp = client.generate_content(prompt)
    return resp.text.strip()


# ── Escrita na planilha ───────────────────────────────────────────────────────

def escrever_mensagens(creds, spreadsheet_id: str, data: str, mensagens: list[dict]):
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.worksheet("Mensagens")

    rows = []
    for m in mensagens:
        rows.append([
            data,
            m.get("fc", ""),
            m.get("setor", ""),
            m.get("atribuicao", ""),
            m.get("turno", ""),
            m.get("matricula", ""),
            m.get("desconto", ""),
            m.get("ideia_central", ""),
            m.get("mensagem", ""),
        ])

    ws.append_rows(rows, value_input_option="USER_ENTERED")
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    print(f"OK: {len(rows)} mensagens escritas na aba 'Mensagens'")
    print(f"LINK={url}")
    return url


# ── PDFs ──────────────────────────────────────────────────────────────────────

def _hex_to_color(hex_str: str):
    h = hex_str.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return colors.Color(r / 255, g / 255, b / 255)


def gerar_pdf_kpi(mensagens: list[dict], data: str, output_path: str):
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    verde = _hex_to_color(COR_VERDE_ESCURO)

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=colors.white, fontSize=14, spaceAfter=4)
    setor_style = ParagraphStyle("setor", parent=styles["Normal"],
                                 textColor=verde, fontName="Helvetica-Bold", fontSize=10)
    msg_style   = ParagraphStyle("msg", parent=styles["Normal"], fontSize=9, leading=13)
    tag_style   = ParagraphStyle("tag", parent=styles["Normal"],
                                 textColor=colors.grey, fontSize=8)

    story = []

    # Cabeçalho principal
    header_data = [[Paragraph(f"Mensagens de Performance FC — {data}", title_style)]]
    header_table = Table(header_data, colWidths=[17*cm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), verde),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4*cm))

    fcs = sorted(set(m["fc"] for m in mensagens))
    for fc in fcs:
        fc_msgs = [m for m in mensagens if m["fc"] == fc]

        fc_header = Table([[Paragraph(fc, title_style)]], colWidths=[17*cm])
        fc_header.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), verde),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(fc_header)
        story.append(Spacer(1, 0.2*cm))

        for m in fc_msgs:
            setor_txt = m.get("setor", "")
            tags = []
            if m.get("atribuicao") and m["atribuicao"] not in ("TODAS", ""):
                tags.append(m["atribuicao"])
            if m.get("turno") and m["turno"] not in ("TODOS", ""):
                tags.append(m["turno"])
            if m.get("desconto"):
                tags.append(m["desconto"])
            tags_txt = " | ".join(tags) if tags else ""

            card_content = [
                [Paragraph(setor_txt, setor_style)],
            ]
            if tags_txt:
                card_content.append([Paragraph(tags_txt, tag_style)])
            card_content.append([Paragraph(m.get("mensagem", ""), msg_style)])

            card = Table([[row[0]] for row in card_content], colWidths=[16*cm])
            card.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f5f5f5")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]))
            story.append(card)
            story.append(Spacer(1, 0.25*cm))

        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    print(f"PDF KPI: {output_path}")


def gerar_pdf_vistoria(vistoria: list[dict], output_path: str):
    entries = [v for v in vistoria if v.get("MATRÍCULA") is not None]
    if not entries:
        print("Vistoria: nenhuma entrada válida, PDF não gerado.")
        return

    periodo = entries[0].get("PERÍODO", "")

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    laranja = _hex_to_color(COR_LARANJA)

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=colors.white, fontSize=13)
    mat_style   = ParagraphStyle("mat", parent=styles["Normal"],
                                 fontName="Helvetica-Bold", fontSize=10)
    msg_style   = ParagraphStyle("msg", parent=styles["Normal"], fontSize=9, leading=13)

    story = []

    header_data = [[Paragraph(f"Mensagens de Vistoria Picking — {periodo}", title_style)]]
    header_table = Table(header_data, colWidths=[17*cm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), laranja),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4*cm))

    fcs = sorted(set(str(v.get("FC", "")) for v in entries if v.get("FC")))
    total = 0
    for fc in fcs:
        fc_entries = [v for v in entries if str(v.get("FC", "")) == fc]
        story.append(Paragraph(f"{fc} — {len(fc_entries)} colaboradores",
                               ParagraphStyle("fch", parent=styles["Heading2"],
                                              textColor=laranja)))
        story.append(Spacer(1, 0.2*cm))

        for i, v in enumerate(fc_entries, 1):
            mat_raw = v.get("MATRÍCULA")
            try:
                mat = str(int(float(mat_raw)))
            except (ValueError, TypeError):
                mat = str(mat_raw)

            pct_raw = v.get("% DESCONTO")
            try:
                pct = int(float(pct_raw) * 100)
                pct_str = f"{pct}%"
            except (ValueError, TypeError):
                pct_str = str(pct_raw)

            mensagem = v.get("MENSAGEM", "")
            zerado = pct_str == "0%"

            bg = colors.HexColor("#fef2f2") if zerado else colors.HexColor("#f9f9f9")
            border_color = colors.red if zerado else colors.lightgrey

            card_rows = [
                [Paragraph(f"{i}. Matrícula {mat} | {pct_str}", mat_style)],
                [Paragraph(mensagem, msg_style)],
            ]
            card = Table([[r[0]] for r in card_rows], colWidths=[16*cm])
            card.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg),
                ("BOX", (0, 0), (-1, -1), 1, border_color),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(card)
            story.append(Spacer(1, 0.2*cm))
            total += 1

        story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(f"Total: {total} colaboradores", styles["Normal"]))
    doc.build(story)
    print(f"PDF Vistoria: {output_path}")


def ler_fiscais_picking(wb, data_quinta: str) -> list[dict]:
    """Lê aba 'Fiscais de Picking', filtrando pelo período da semana."""
    if "Fiscais de Picking" not in wb.sheetnames:
        return []
    ws = wb["Fiscais de Picking"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    quinta = datetime.strptime(data_quinta, "%d/%m/%Y").date()
    sexta = quinta - timedelta(days=6)
    periodo_str = f"{sexta.strftime('%d.%m')}-{quinta.strftime('%d.%m')}"

    result = []
    for row in rows[1:]:
        parsed = [extract_val(c) for c in row]
        if not any(v for v in parsed):
            continue
        if len(parsed) < 10:
            continue
        matricula, _, operador, setor, atribuicao, turno, fc, data_apur, pct_desconto, mensagem = parsed[:10]
        if data_apur and periodo_str in str(data_apur):
            try:
                mat = str(int(float(matricula))) if matricula is not None else None
            except (ValueError, TypeError):
                mat = str(matricula) if matricula is not None else None
            if not mat:
                continue
            result.append({
                "matricula": mat,
                "fc": str(fc) if fc else "",
                "setor": str(setor) if setor else "",
                "atribuicao": str(atribuicao) if atribuicao else "",
                "turno": str(turno) if turno else "",
                "pct_desconto": pct_desconto,
                "mensagem": str(mensagem) if mensagem else "",
            })
    return result


# ── Atualização do SQL ────────────────────────────────────────────────────────

def _sql_mats(mats: list) -> str:
    return ", ".join(f"'{m}'" for m in mats)


def _desconto_to_mult(desconto_str) -> str:
    """'-25%'→'0.75', '0%'→'0.0', float -0.5→'0.5'"""
    if desconto_str is None:
        return "1.0"
    s = str(desconto_str).strip()
    try:
        if "%" in s:
            pct = float(s.replace("%", "").replace("+", ""))
            mult = 1.0 + pct / 100.0
        else:
            mult = 1.0 + float(s)
        return str(round(max(0.0, mult), 4))
    except ValueError:
        return "1.0"


def _setor_to_sql_cond(setor: str, atribuicao: str, turno: str, fc: str) -> str:
    s = setor.upper().strip()
    a = (atribuicao or "").upper().strip()
    t = (turno or "").upper().strip()
    parts = []

    if "PRÉ" in s or "PRE" in s:
        if "EXPED" in s:
            parts.append("(SETOR_ORIGINAL LIKE '%PRÉ%EXPED%' OR SETOR_ORIGINAL LIKE '%PRE%EXPED%')")
        else:
            parts.append(f"SETOR_ORIGINAL LIKE '%{s}%'")
    elif "REPOSIÇÃO" in s or "REPOSICAO" in s or "REPOSICÃO" in s:
        parts.append("SETOR_ORIGINAL IN ('REPOSIÇÃO', 'REPOSICAO')")
    elif "RECEBIMENTO" in s:
        parts.append("SETOR_ORIGINAL LIKE '%RECEBIMENTO%'")
    elif "EXPEDIÇÃO" in s or "EXPEDICAO" in s:
        parts.append("SETOR_ORIGINAL IN ('EXPEDIÇÃO', 'EXPEDICAO')")
    elif "PICKING" in s:
        parts.append("SETOR_ORIGINAL = 'PICKING'")
    elif "PACKING" in s:
        parts.append("SETOR_ORIGINAL = 'PACKING'")
    elif "FRACIONAMENTO" in s or "FRAC" in s:
        parts.append("SETOR_ORIGINAL LIKE '%FRACIONAMENTO%'")
    else:
        parts.append(f"UPPER(TRIM(SETOR_ORIGINAL)) = '{s}'")

    if "FRESH" in s:
        parts.append("AREA = 'FRESH'")
    elif "MERCEARIA" in s:
        parts.append("AREA = 'MERCEARIA'")

    parts.append(f"FC = '{fc}'")

    if a and a not in ("TODAS", "TODOS", ""):
        for kw in ("FLV", "CONGELADO", "REFRIGERADO"):
            if kw in a:
                parts.append(f"ATRIBUICAO_ORIGINAL LIKE '%{kw}%'")
                break
        else:
            parts.append(f"ATRIBUICAO_ORIGINAL LIKE '%{a}%'")

    if t and t not in ("TODOS", "TODAS", ""):
        turnos = [x.strip() for x in t.replace("/", ",").split(",")]
        turno_list = ", ".join(f"'{tu}'" for tu in turnos)
        cond = f"TURNO IN ({turno_list})" if len(turnos) > 1 else f"TURNO = {turno_list}"
        parts.append(cond)

    return " AND ".join(parts)


def _sql_str(s: str) -> str:
    """Sanitiza string para uso em literal SQL: remove newlines, escapa aspas simples."""
    return re.sub(r"\s*\n\s*", " ", str(s)).strip().replace("'", "''")


def _replace_sql_section(sql: str, section: str, content: str) -> str:
    pattern = (
        rf"([ \t]*)-- \[AUTO:{re.escape(section)}\]\n"
        rf".*?"
        rf"([ \t]*)-- \[/AUTO:{re.escape(section)}\]"
    )
    def replacer(m):
        indent = m.group(1)
        # Re-indent each content line with the block's own indent
        indented = ""
        for line in content.splitlines(keepends=True):
            stripped = line.lstrip()
            indented += (indent + stripped) if stripped else line
        return f"{indent}-- [AUTO:{section}]\n{indented}{indent}-- [/AUTO:{section}]"
    return re.sub(pattern, replacer, sql, flags=re.DOTALL)


def _to_mat(val) -> str:
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str(val) if val is not None else ""


def _is_zero(desconto) -> bool:
    if desconto is None:
        return False
    s = str(desconto).strip().replace("%", "").replace("+", "")
    try:
        return float(s) == 0.0
    except ValueError:
        return False


def atualizar_sql_kpis(
    sql_path: str,
    data: str,
    kpis_ind: list,
    vistoria: list,
    fiscais: list,
    visao_por_fc: dict,
    mensagens: list,
    ge_rows: list,
):
    with open(sql_path, encoding="utf-8") as f:
        sql = f.read()

    hoje_str = datetime.strptime(data, "%d/%m/%Y").strftime("%d/%m/%Y")
    sql = re.sub(r"-- Última atualização: [\d/]+", f"-- Última atualização: {hoje_str}", sql)

    # ── 1. ind-zerados-mult (MULT_MATRICULA — KPIs Individuais 0%) ────────────
    zerado_mats = sorted({_to_mat(d["MATRÍCULA"]) for d in kpis_ind
                          if d.get("MATRÍCULA") and _is_zero(d.get("DESCONTO"))})
    ind_zerados_mult = (f"WHEN MATRICULA IN ({_sql_mats(zerado_mats)}) THEN 0.0\n"
                        if zerado_mats else "")
    sql = _replace_sql_section(sql, "ind-zerados-mult", ind_zerados_mult)

    # ── 2. ind-parcial-mult (MULT_MATRICULA — KPIs Individuais com desconto) ──
    parcial_groups: dict = {}
    for d in kpis_ind:
        if not d.get("MATRÍCULA") or _is_zero(d.get("DESCONTO")):
            continue
        mat = _to_mat(d["MATRÍCULA"])
        mult = _desconto_to_mult(d.get("DESCONTO"))
        parcial_groups.setdefault(mult, []).append(mat)
    ind_parcial_lines = [
        f"WHEN MATRICULA IN ({_sql_mats(mats)}) THEN {mult}"
        for mult, mats in parcial_groups.items()
    ]
    sql = _replace_sql_section(
        sql, "ind-parcial-mult",
        "\n".join(ind_parcial_lines) + "\n" if ind_parcial_lines else "",
    )

    # ── 3. ind-zerados-setor-neut (neutralizador em MULT_SETOR) ──────────────
    sql = _replace_sql_section(
        sql, "ind-zerados-setor-neut",
        f"WHEN MATRICULA IN ({_sql_mats(zerado_mats)}) THEN 1.0\n" if zerado_mats else "",
    )

    # ── 4. fiscais-picking-mult (MULT_MATRICULA — Vistoria + Fiscais) ─────────
    # Agrupa Vistoria por (mult, mensagem) para compactar matrículas iguais
    picking_groups: dict = {}
    for v in vistoria:
        mat_raw = v.get("MATRÍCULA")
        if mat_raw is None:
            continue
        try:
            mat = str(int(float(mat_raw)))
        except (ValueError, TypeError):
            continue
        pct_raw = v.get("% DESCONTO")
        try:
            mult = str(round(max(0.0, 1.0 + float(pct_raw)), 4)).rstrip("0").rstrip(".")
            if "." not in mult:
                mult += ".0"
        except (ValueError, TypeError):
            continue
        msg = str(v.get("MENSAGEM", "") or "").replace("'", "''")
        picking_groups.setdefault((mult, msg), []).append(mat)

    for f in fiscais:
        mat = f.get("matricula")
        if not mat:
            continue
        mult = _desconto_to_mult(f.get("pct_desconto"))
        msg = (f.get("mensagem") or "").replace("'", "''")
        picking_groups.setdefault((mult, msg), []).append(str(mat))

    picking_mult_lines = [
        f"WHEN MATRICULA IN ({_sql_mats(mats)}) THEN {mult}"
        for (mult, msg), mats in picking_groups.items()
    ]
    sql = _replace_sql_section(
        sql, "fiscais-picking-mult",
        "\n".join(picking_mult_lines) + "\n" if picking_mult_lines else "",
    )

    # ── 5. setoriais-mult (MULT_SETOR, por setor/FC) ──────────────────────────
    setor_mult_lines = []
    for fc, visao in visao_por_fc.items():
        for row in visao:
            setor = row.get("SETOR", "")
            if not setor:
                continue
            desconto = row.get("DESCONTO", "")
            if not desconto or str(desconto).strip() in ("", "0%", "0"):
                continue
            mult = _desconto_to_mult(desconto)
            atrib = row.get("ATRIBUIÇÃO", "") or ""
            turno = row.get("TURNO", "") or ""
            cond = _setor_to_sql_cond(setor, atrib, turno, fc)
            setor_mult_lines.append(f"WHEN {cond} THEN {mult}")
    sql = _replace_sql_section(
        sql, "setoriais-mult",
        "\n".join(setor_mult_lines) + "\n" if setor_mult_lines else "",
    )

    # ── 6. ind-zerados-obs (OBSERVACAO_KPI — KPIs Individuais 0%) ────────────
    ind_obs_lines = []
    for d in kpis_ind:
        if not d.get("MATRÍCULA") or not _is_zero(d.get("DESCONTO")):
            continue
        mat = _to_mat(d["MATRÍCULA"])
        motivo = _sql_str(d.get("OBS (MOTIVO)", "") or "")
        msg = f"VALOR DA BONIFICAÇÃO ZERADO. {motivo}" if motivo else "VALOR DA BONIFICAÇÃO ZERADO."
        ind_obs_lines.append(f"WHEN MATRICULA IN ('{mat}')\n  THEN '{msg}'")
    sql = _replace_sql_section(
        sql, "ind-zerados-obs",
        "\n".join(ind_obs_lines) + "\n" if ind_obs_lines else "",
    )

    # ── 6b. ind-parcial-obs (OBSERVACAO_KPI — KPIs Individuais parciais) ──────
    ind_parcial_obs_lines = []
    for d in kpis_ind:
        if not d.get("MATRÍCULA") or _is_zero(d.get("DESCONTO")):
            continue
        mat = _to_mat(d["MATRÍCULA"])
        motivo = _sql_str(d.get("OBS (MOTIVO)", "") or "")
        desconto_str = str(d.get("DESCONTO", "")).strip()
        msg = f"{desconto_str} na Bonificação. {motivo}" if motivo else f"{desconto_str} na Bonificação."
        ind_parcial_obs_lines.append(f"WHEN MATRICULA IN ('{mat}')\n  THEN '{msg}'")
    sql = _replace_sql_section(
        sql, "ind-parcial-obs",
        "\n".join(ind_parcial_obs_lines) + "\n" if ind_parcial_obs_lines else "",
    )

    # ── 7. fiscais-picking-obs (Vistoria + Fiscais, agrupado por mensagem) ────
    fiscais_obs_lines = []
    for (mult, msg), mats in picking_groups.items():
        if msg:
            fiscais_obs_lines.append(f"WHEN MATRICULA IN ({_sql_mats(mats)})\n  THEN '{_sql_str(msg)}'")
    sql = _replace_sql_section(
        sql, "fiscais-picking-obs",
        "\n".join(fiscais_obs_lines) + "\n" if fiscais_obs_lines else "",
    )

    # ── 8. ge-obs ─────────────────────────────────────────────────────────────
    ge_obs_lines = []
    msg_to_mats: dict = {}
    for g in ge_rows:
        mat = g.get("matricula")
        msg_raw = g.get("mensagem") or ""
        if not mat or not msg_raw:
            continue
        msg_to_mats.setdefault(_sql_str(msg_raw), []).append(mat)
    for msg, mats in msg_to_mats.items():
        ge_obs_lines.append(f"WHEN MATRICULA IN ({_sql_mats(mats)})\n  THEN '{msg}'")
    sql = _replace_sql_section(
        sql, "ge-obs",
        "\n".join(ge_obs_lines) + "\n" if ge_obs_lines else "",
    )

    # ── 9. setoriais-obs ──────────────────────────────────────────────────────
    setor_obs_lines = []
    for fc, visao in visao_por_fc.items():
        for row in visao:
            setor = row.get("SETOR", "")
            if not setor:
                continue
            desconto = row.get("DESCONTO", "")
            if not desconto or str(desconto).strip() in ("", "0%", "0"):
                continue
            atrib = row.get("ATRIBUIÇÃO", "") or ""
            turno = row.get("TURNO", "") or ""
            cond = _setor_to_sql_cond(setor, atrib, turno, fc)
            msg = ""
            for m in mensagens:
                if (m.get("fc") == fc and
                        m.get("setor", "").upper() == setor.upper() and
                        m.get("atribuicao", "").upper() in (atrib.upper(), "TODAS", "") and
                        m.get("turno", "").upper() in (turno.upper(), "TODOS", "")):
                    msg = m.get("mensagem", "")
                    break
            if msg:
                setor_obs_lines.append(
                    f"WHEN {cond}\n  THEN '{_sql_str(msg)}'"
                )
    sql = _replace_sql_section(
        sql, "setoriais-obs",
        "\n".join(setor_obs_lines) + "\n" if setor_obs_lines else "",
    )

    with open(sql_path, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"SQL atualizado: {sql_path}")


def gerar_pdf_detratores(detratores: list[dict], data: str, output_path: str):
    if not detratores:
        print("Detratores: nenhuma entrada, PDF não gerado.")
        return

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    cor_header = colors.HexColor("#1a3a5c")

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=colors.white, fontSize=13)
    cell_style  = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=11)
    hdr_style   = ParagraphStyle("hdr", parent=styles["Normal"],
                                 textColor=colors.white, fontName="Helvetica-Bold", fontSize=8)

    story = []

    header_data = [[Paragraph(f"Detratores Individuais — {data}", title_style)]]
    header_table = Table(header_data, colWidths=[17*cm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), cor_header),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4*cm))

    col_widths = [2.2*cm, 1.8*cm, 4.5*cm, 1.5*cm, 1.8*cm, 2.0*cm, 3.2*cm]
    col_headers = ["DATA", "MATRÍCULA", "NOME", "FC", "TURNO", "DESCONTO", "MOTIVO"]

    table_data = [[Paragraph(h, hdr_style) for h in col_headers]]
    for d in detratores:
        mat_raw = d.get("MATRÍCULA")
        try:
            mat = str(int(float(mat_raw)))
        except (ValueError, TypeError):
            mat = str(mat_raw) if mat_raw is not None else ""

        pct_raw = d.get("DESCONTO", "")
        pct_str = str(pct_raw) if pct_raw is not None else ""

        table_data.append([
            Paragraph(str(d.get("DATA (QUINTA)", "")), cell_style),
            Paragraph(mat, cell_style),
            Paragraph(str(d.get("NOME", "")), cell_style),
            Paragraph(str(d.get("FC", "")), cell_style),
            Paragraph(str(d.get("TURNO", "")), cell_style),
            Paragraph(pct_str, cell_style),
            Paragraph(str(d.get("OBS (MOTIVO)", "") or ""), cell_style),
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), cor_header),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Total: {len(detratores)} detratores", styles["Normal"]))

    doc.build(story)
    print(f"PDF Detratores: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", required=False, help="Data no formato DD/MM/AAAA")
    parser.add_argument("--output-dir", default=".", help="Diretório de saída dos PDFs")
    parser.add_argument("--sql-path", required=False, help="Caminho para 00_kpis_operacao.sql a atualizar")
    args = parser.parse_args()

    if args.data:
        data = args.data
    else:
        hoje = datetime.today().date()
        # 3 = quinta-feira (weekday: seg=0 … dom=6)
        dias_ate_quinta = (hoje.weekday() - 3) % 7
        data = (hoje - timedelta(days=dias_ate_quinta)).strftime("%d/%m/%Y")
    data_fmt = data.replace("/", "")  # para nome de arquivo
    output_dir = args.output_dir
    os.makedirs(output_dir, exist_ok=True)

    print(f"Iniciando geração de mensagens FC — {data}")

    creds = get_credentials()
    genai.configure(api_key=os.environ["GEMINI_API_KEY"])
    gemini_client = genai.GenerativeModel("gemini-2.5-pro")

    print("Baixando planilha Dashboard FC...")
    wb_fc = download_xlsx(SPREADSHEET_ID_FC, creds)

    print("Baixando planilha Performance inventário (GE)...")
    wb_ge = download_xlsx(SPREADSHEET_ID_GE, creds)

    mensagens_finais = []

    # ── Mensagens setoriais (Visão FC + Resumo FC) ────────────────────────────
    for fc in ["FC1", "FC2", "FC3"]:
        resumo = ler_resumo_fc(wb_fc, fc)
        visao  = ler_visao_fc(wb_fc, fc)

        for config in visao:
            setor = config.get("SETOR", "")
            if not setor:
                continue
            print(f"Gerando mensagem: {fc} / {setor}")
            mensagem = gerar_mensagem_setor(gemini_client, fc, config, resumo)
            mensagens_finais.append({
                "fc": fc,
                "setor": setor,
                "atribuicao": config.get("ATRIBUIÇÃO", "") or "TODAS",
                "turno": config.get("TURNO", "") or "TODOS",
                "matricula": config.get("MATRÍCULA", "") or "",
                "desconto": config.get("DESCONTO", "") or "",
                "ideia_central": config.get("IDEIA CENTRAL", "") or "",
                "mensagem": mensagem,
            })

    # ── Fiscais de Picking ────────────────────────────────────────────────────
    fiscais = ler_fiscais_picking(wb_fc, data)
    print(f"Fiscais de Picking: {len(fiscais)} entradas para o período")

    # ── Mensagens GE ──────────────────────────────────────────────────────────
    ge_rows = ler_ge(wb_ge)
    print(f"GE: {len(ge_rows)} linhas com JUSTIFICATIVA preenchida")
    for item in ge_rows:
        print(f"Gerando mensagem GE: {item['fc']} / matrícula {item['matricula']}")
        mensagem = gerar_mensagem_ge(gemini_client, item)
        mensagens_finais.append({
            "fc": item["fc"],
            "setor": "GESTÃO DE ESTOQUE",
            "atribuicao": "INVENTÁRIO",
            "turno": "",
            "matricula": item["matricula"] or "",
            "desconto": "",
            "ideia_central": item["justificativa"],
            "mensagem": mensagem,
        })

    # ── Escrever na planilha ──────────────────────────────────────────────────
    escrever_mensagens(creds, SPREADSHEET_ID_FC, data, mensagens_finais)

    # ── PDFs ──────────────────────────────────────────────────────────────────
    msgs_setoriais = [m for m in mensagens_finais if m["setor"] != "GESTÃO DE ESTOQUE"]
    pdf_kpi = os.path.join(output_dir, f"mensagens_kpi_fc_{data_fmt}.pdf")
    gerar_pdf_kpi(msgs_setoriais, data, pdf_kpi)

    vistoria = ler_vistoria_semana(wb_fc)
    pdf_vistoria = os.path.join(output_dir, f"vistoria_picking_{data_fmt}.pdf")
    gerar_pdf_vistoria(vistoria, pdf_vistoria)

    detratores = ler_kpis_individuais(wb_fc, data)
    if detratores:
        pdf_detratores = os.path.join(output_dir, f"detratores_individuais_{data_fmt}.pdf")
        gerar_pdf_detratores(detratores, data, pdf_detratores)
    else:
        print("KPIs Individuais: nenhum detrator para a data informada.")

    # ── Atualizar SQL ──────────────────────────────────────────────────────────
    if args.sql_path:
        visao_por_fc = {}
        for fc in ["FC1", "FC2", "FC3"]:
            visao_por_fc[fc] = ler_visao_fc(wb_fc, fc)

        ge_com_msg = []
        for m in mensagens_finais:
            if m.get("setor") == "GESTÃO DE ESTOQUE":
                ge_com_msg.append({
                    "matricula": m.get("matricula"),
                    "mensagem": m.get("mensagem"),
                })

        atualizar_sql_kpis(
            sql_path=args.sql_path,
            data=data,
            kpis_ind=detratores,
            vistoria=vistoria,
            fiscais=fiscais,
            visao_por_fc=visao_por_fc,
            mensagens=mensagens_finais,
            ge_rows=ge_com_msg,
        )

    print(f"\nConcluído. {len(mensagens_finais)} mensagens geradas.")


if __name__ == "__main__":
    main()
